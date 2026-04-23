from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage

logger = logging.getLogger(__name__)

TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "instruction_template.xlsx"
START_ROW = 8


class InstructionSheet:
    def __init__(self, save_dir: Path, keepa_api_key: str) -> None:
        self._save_dir = save_dir
        self._keepa_api_key = keepa_api_key

    def create(self, data: list[Any]) -> Path:
        rows = self._extract_rows(data)
        plan_name = self._generate_plan_name(data)

        wb = load_workbook(str(TEMPLATE_PATH))
        ws = wb.active
        self._write_row_data(ws, rows)

        self._save_dir.mkdir(parents=True, exist_ok=True)
        file_path = self._save_dir / f"{plan_name}.xlsx"
        wb.save(str(file_path))
        logger.info("指示書保存: %s", file_path)
        return file_path

    def _extract_rows(self, data: list[Any]) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        for row in data:
            fnsku = str(row.get("FNSKU") or "").strip()
            asin = str(row.get("ASIN") or "").strip()
            quantity = str(row.get("購入数") or "").strip()
            remarks = str(row.get("備考") or "").strip()
            order_number = str(row.get("注文番号") or "").strip()
            if fnsku:
                rows.append({"fnsku": fnsku, "asin": asin, "quantity": quantity, "remarks": remarks, "order_number": order_number})
        return rows

    def _generate_plan_name(self, data: list[Any]) -> str:
        now = datetime.now()
        date_str = f"{now.month:02d}{now.day:02d}"
        try:
            category = str(data[0].get("納品分類") or "").strip() if data else ""
        except Exception:
            category = ""
        return f"{date_str}{category}指示書"

    def _write_row_data(self, ws: Any, rows: list[dict[str, str]]) -> None:
        for i, row_data in enumerate(rows):
            row_num = START_ROW + i
            # B列: FNSKU, C列: ASIN, D列: 数量（テンプレートの列順に合わせる）
            ws.cell(row=row_num, column=2, value=row_data["fnsku"])
            ws.cell(row=row_num, column=3, value=row_data["asin"])
            ws.cell(row=row_num, column=4, value=int(row_data["quantity"]) if row_data["quantity"] else 0)
            ws.cell(row=row_num, column=5, value=row_data["remarks"])
            ws.cell(row=row_num, column=6, value=row_data["order_number"])

            image_url = self._get_product_image(row_data["asin"])
            if image_url:
                try:
                    img_response = httpx.get(image_url, timeout=10.0)
                    if img_response.status_code == 200:
                        import tempfile
                        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
                            tmp.write(img_response.content)
                            tmp_path = tmp.name
                        img = XlImage(tmp_path)
                        img.width = 75
                        img.height = 75
                        ws.add_image(img, f"A{row_num}")
                except Exception as e:
                    logger.warning("画像挿入エラー (%s): %s", row_data["asin"], e)

    def _get_product_image(self, asin: str) -> str | None:
        if not asin or not self._keepa_api_key:
            return None
        try:
            url = f"https://api.keepa.com/product?key={self._keepa_api_key}&domain=5&asin={asin}"
            response = httpx.get(url, timeout=30.0)
            data = response.json()
            products = data.get("products", [])
            if not products:
                return None
            product = products[0]
            images_csv = product.get("imagesCSV", "")
            if images_csv:
                first_image = images_csv.split(",")[0]
                return f"https://images-na.ssl-images-amazon.com/images/I/{first_image}._SL100_.jpg"
            images = product.get("images", [])
            if images:
                image_id = images[0].get("m") or images[0].get("l", "")
                if image_id:
                    return f"https://images-na.ssl-images-amazon.com/images/I/{image_id}"
            return None
        except Exception as e:
            logger.warning("Keepa画像取得エラー (%s): %s", asin, e)
            return None
