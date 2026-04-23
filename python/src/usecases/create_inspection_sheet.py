from __future__ import annotations

import io
import logging
import re
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage

from shared.config import AppConfig
from infrastructure.spreadsheet.base_sheets_repository import BaseSheetsRepository
from infrastructure.spreadsheet.inspection_master_repo import InspectionMasterRepo

logger = logging.getLogger(__name__)

INSPECTION_TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "infrastructure" / "spreadsheet" / "templates" / "inspection_template.xlsx"
START_ROW = 9
ROW_STEP = 5


def create_inspection_sheet_if_needed(
    config: AppConfig, repo: BaseSheetsRepository, rows: list[Any], category: str
) -> Path | None:
    if not config.inspection_master_sheet_id:
        return None

    master_repo = InspectionMasterRepo(repo, config.inspection_master_sheet_id, config.inspection_master_sheet_gid)
    catalog = master_repo.load()

    matched = _collect_matched_items(rows, catalog)
    if not matched:
        logger.info("検品マスタに該当ASINなし -> スキップ")
        return None

    logger.info("[検品] 検品シート書き込み対象件数=%d", len(matched))

    wb = load_workbook(str(INSPECTION_TEMPLATE_PATH))
    ws = wb.active
    _write_inspection_data(ws, matched, config.keepa_api_key)
    _append_detail_instruction_sheets(wb, matched, repo)

    save_dir = Path(config.detail_inspection_dir)
    save_dir.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%m%d")
    file_path = save_dir / f"{date_str}{category}検品指示書.xlsx"
    wb.save(str(file_path))
    logger.info("検品指示書保存: %s", file_path)
    return file_path


def _collect_matched_items(rows: list[Any], catalog: Any) -> list[dict[str, Any]]:
    matched: list[dict[str, Any]] = []
    for r in rows:
        try:
            asin = str(r.get("ASIN") or "").strip()
        except Exception:
            continue
        if not asin:
            continue
        master_item = catalog.get(asin)
        if not master_item:
            continue
        try:
            order_no = str(r.get("注文番号") or "").strip()
        except Exception:
            order_no = ""
        try:
            quantity = int(r.get("購入数") or 0)
        except Exception:
            quantity = 0

        matched.append({
            "asin": asin,
            "order_no": order_no,
            "product_name": master_item.product_name,
            "quantity": quantity,
            "inspection_point": master_item.inspection_point,
            "detail_instruction_url": master_item.detail_instruction_url,
        })
    return matched


def _write_inspection_data(ws: Any, matched: list[dict[str, Any]], keepa_api_key: str) -> None:
    _unmerge_data_cells(ws, len(matched))

    for i, item in enumerate(matched):
        target_row = START_ROW + (i * ROW_STEP)
        ws.cell(row=target_row, column=3, value=item["order_no"])
        ws.cell(row=target_row, column=4, value=item["product_name"])
        ws.cell(row=target_row, column=7, value=item["quantity"])
        ws.cell(row=target_row, column=8, value=item["inspection_point"])

        image_url = _get_product_image_url(item["asin"], keepa_api_key)
        if image_url:
            try:
                img_response = httpx.get(image_url, timeout=10.0)
                if img_response.status_code == 200:
                    import tempfile
                    with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
                        tmp.write(img_response.content)
                        tmp_path = tmp.name
                    img = XlImage(tmp_path)
                    img.width = 75
                    img.height = 75
                    ws.add_image(img, f"E{target_row}")
            except Exception as e:
                logger.warning("画像挿入エラー (%s): %s", item["asin"], e)


def _unmerge_data_cells(ws: Any, item_count: int) -> None:
    rows_to_unmerge: set[int] = set()
    for i in range(item_count):
        target_row = START_ROW + (i * ROW_STEP)
        for r in range(target_row, target_row + ROW_STEP):
            rows_to_unmerge.add(r)

    ranges_to_unmerge: list[str] = []
    for merged_range in list(ws.merged_cells.ranges):
        if any(r in rows_to_unmerge for r in range(merged_range.min_row, merged_range.max_row + 1)):
            ranges_to_unmerge.append(str(merged_range))

    for r in ranges_to_unmerge:
        try:
            ws.unmerge_cells(r)
        except Exception:
            pass


def _append_detail_instruction_sheets(
    wb: Any, matched: list[dict[str, Any]], repo: BaseSheetsRepository
) -> None:
    items_with_url = [it for it in matched if it.get("detail_instruction_url")]
    if not items_with_url:
        logger.info("[検品] 詳細指示書URLなし -> シート追加スキップ")
        return

    logger.info("[検品] 詳細指示書URLあり件数=%d -> シートコピー開始", len(items_with_url))

    for item in items_with_url:
        parsed = _parse_spreadsheet_url(item["detail_instruction_url"])
        if not parsed or not parsed.get("spreadsheet_id"):
            logger.warning("[検品] 詳細指示書URLを解析できません: ASIN=%s, url=%s", item["asin"], item["detail_instruction_url"])
            continue

        try:
            src_ss = repo.client.open_by_key(parsed["spreadsheet_id"])

            # xlsxとしてエクスポートし、書式ごとコピー
            xlsx_data = src_ss.export(format="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            src_wb = load_workbook(io.BytesIO(xlsx_data))

            # 対象シートを特定
            if parsed.get("gid"):
                src_ws = None
                for ws_obj in src_ss.worksheets():
                    if str(ws_obj.id) == str(parsed["gid"]):
                        src_ws_title = ws_obj.title
                        src_ws = src_wb[src_ws_title] if src_ws_title in src_wb.sheetnames else None
                        break
                if not src_ws:
                    src_ws = src_wb.active
            else:
                src_ws = src_wb.active

            sheet_name = f"検品詳細_{item['product_name']}"[:31]
            sheet_name = _make_unique_sheet_name(wb, sheet_name)
            new_ws = wb.create_sheet(title=sheet_name)

            _copy_worksheet(src_ws, new_ws)

            logger.info("[検品] 詳細指示書シート追加: ASIN=%s, name=%s", item["asin"], sheet_name)

        except Exception as e:
            logger.warning("[検品] 詳細指示書シート追加に失敗: ASIN=%s, error=%s", item["asin"], e)


def _copy_worksheet(src_ws: Any, dst_ws: Any) -> None:
    # 列幅をコピー
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        dst_ws.column_dimensions[col_letter].hidden = dim.hidden

    # 行高さをコピー
    for row_num, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = dim.height
        dst_ws.row_dimensions[row_num].hidden = dim.hidden

    # セル結合をコピー
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    # 画像をコピー
    for image in src_ws._images:
        new_image = XlImage(image.ref)
        new_image.anchor = copy(image.anchor)
        new_image.width = image.width
        new_image.height = image.height
        dst_ws.add_image(new_image)

    # セルの値と書式をコピー
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)


def _parse_spreadsheet_url(url: str) -> dict[str, str] | None:
    s = str(url or "").strip()
    if not s:
        return None
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", s)
    spreadsheet_id = m.group(1) if m else None
    if not spreadsheet_id:
        return None
    result: dict[str, str] = {"spreadsheet_id": spreadsheet_id}
    gid_match = re.search(r"[?#&]gid=(\d+)", s)
    if gid_match:
        result["gid"] = gid_match.group(1)
    return result


def _make_unique_sheet_name(wb: Any, base_name: str) -> str:
    existing = {ws.title for ws in wb.worksheets}
    if base_name not in existing:
        return base_name
    for i in range(2, 100):
        name = f"{base_name}_{i}"[:31]
        if name not in existing:
            return name
    return f"{base_name[:20]}_{int(datetime.now().timestamp())}"[:31]


def _get_product_image_url(asin: str, keepa_api_key: str) -> str | None:
    if not keepa_api_key:
        return None
    try:
        url = f"https://api.keepa.com/product?key={keepa_api_key}&domain=5&asin={asin}"
        response = httpx.get(url, timeout=30.0)
        data = response.json()
        products = data.get("products", [])
        if not products:
            return None
        images_csv = products[0].get("imagesCSV", "")
        if not images_csv:
            return None
        first_image = images_csv.split(",")[0]
        return f"https://images-na.ssl-images-amazon.com/images/I/{first_image}._SL100_.jpg"
    except Exception as e:
        logger.warning("Keepa画像取得エラー (%s): %s", asin, e)
        return None
